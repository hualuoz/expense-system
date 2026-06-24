#!/usr/bin/env python3
"""异步任务模块 — OCR + Excel导出"""

import threading
import json
import io
import os
import sys

WEB_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, WEB_DIR)  # web/validate_expense.py, web/file_parser.py

from app.dao.db import new_id
from app.dao.models import AsyncTaskDAO


def run_async_task(task_id, task_type, task_func, *args):
    def worker():
        try:
            result = task_func(*args)
            AsyncTaskDAO.update_result(task_id, "COMPLETED", json.dumps(result, ensure_ascii=False, default=str))
        except Exception as e:
            AsyncTaskDAO.update_result(task_id, "FAILED", "", str(e))
    t = threading.Thread(target=worker, daemon=True)
    t.start()


def submit_ocr_task(filename, file_content):
    task_id = new_id()
    AsyncTaskDAO.create(task_id, "OCR", json.dumps({"filename": filename}, ensure_ascii=False))

    def do_ocr():
        from file_parser import parse_file
        return parse_file(filename, file_content)

    run_async_task(task_id, "OCR", do_ocr)
    return task_id


def submit_export_task(rid, user_id):
    task_id = new_id()
    AsyncTaskDAO.create(task_id, "EXPORT", json.dumps({"rid": rid, "user_id": user_id}, ensure_ascii=False))

    def do_export():
        from app.dao.models import ReimbursementDAO, ExpenseItemDAO
        from app.dao.db import Status
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        r = ReimbursementDAO.get_by_id(rid)
        if not r:
            raise ValueError("报销单不存在")
        items = ExpenseItemDAO.list_by_reimbursement(rid)

        wb = Workbook()
        ws = wb.active
        ws.title = "报销申请单"
        ws.merge_cells('A1:F1')
        ws['A1'] = '报销申请单'
        ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A3'] = '申请单号:'; ws['B3'] = rid
        ws['D3'] = '申请人:';   ws['E3'] = r['applicant_name']
        ws['A4'] = '所属部门:'; ws['B4'] = r['department']
        ws['D4'] = '状态:';     ws['E4'] = Status.DISPLAY.get(r['status'], r['status'])
        ws['A5'] = '合计金额:'; ws['E5'] = f"¥{r['total_amount']:,.2f}"

        hf = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        hfill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        headers = ['序号', '日期', '费用类别', '金额(元)', '说明', '票据号']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=7, column=i, value=h)
            c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center'); c.border = bd

        for idx, item in enumerate(items):
            row = 8 + idx
            vals = [idx+1, item['date'], item['category'], item['amount'], item['description'], item['receipt_no']]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name='微软雅黑', size=10); c.border = bd
                if col == 4: c.number_format = '#,##0.00'

        ws.column_dimensions['A'].width = 8; ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 20; ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 30; ws.column_dimensions['F'].width = 18

        export_dir = os.path.join(WEB_DIR, "exports")
        os.makedirs(export_dir, exist_ok=True)
        export_path = os.path.join(export_dir, f"报销单_{rid}.xlsx")
        wb.save(export_path)
        return {"filepath": export_path, "filename": f"报销单_{rid}.xlsx"}

    run_async_task(task_id, "EXPORT", do_export)
    return task_id
